import sys
import os
import json
import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.utils import cell

# You can change these to your desired filenames.
def get_default_files():
    """
    If no arguments are passed by command line, then it will run these files by
    default.

    in_file - The input workbook that you want to take information from.
    template_file - The template file of the desired output format.
    out_file - The filename for the new output file. Be cautious of overwriting.
    col_file - The JSON file that contains column mappings from input to output
               template.
    """
    in_file = "PKS Sample.xlsx"
    template_file = "Tracker Template.xlsx"
    out_file = "output.xlsx"
    col_file = "col_config.json"
    return in_file, template_file, out_file, col_file

def read_col_json(col_file):
    """
    Reads a the file of name col_file and loads its JSON. If the file passes some checks, it will return the parsed JSON
    dictionary. Otherwise, it will return an empty dictionary.

    col_file - The name of the column configuration JSON file. This is a file that, at its most basic form, requires two
               top-level keys, "col_keys" and "col_start", each with their own dictionaries. The "col_keys" must contain
               values corresponding to the column names from the input excel as keys that map to column names from the
               template excel. The "col_start" must contain the column names from the template excel as the keys that map
               to the location of the corresponding column header in the template excel. For more details, please refer to
               the user guide.
    """
    # Open file and read JSON
    with open(col_file, 'r') as f:
        jcols = json.load(f)

    # Return the JSON dict if it passes some early checks
    if validate_jcols_keys(jcols, col_file):
        return jcols

    return {}

def validate_jcols_keys(jcols, col_file):
    """
    Validates the col_config.json file. There must be two keys at the top-level: "col_keys" and "col_start". The values in
    col_keys must match the keys of col_start. The cell locations in col_start must adhere to standard Excel coordinate
    convention (for example "A1"). The character (column) component can contain one to three alphabetical letters, and
    there is no validation on the digit component as long as there is at least one digit.

    jcols - The JSON dictionary read from col_file.
    col_file - The name of the column configuration JSON file.
    """
    # Check length
    if len(jcols) != 2:
        print("Error: Please verify the top level of \"{}\" contains two dictionaries, \"col_keys\" and \"col_start\".".format(
             col_file))
        return False

    # Check top keys are col_keys and col_start
    top_keys = jcols.keys()
    if ("col_keys" not in top_keys) | ("col_start" not in top_keys):
        print("Error: Please verify the top level of \"{}\" contains the keys \"col_keys\" and \"col_start\".".format(
            col_file))
        return False

    # Check col_keys mapped template cols match col_start template cols
    key_outmaps = []
    for v in jcols['col_keys'].values():
        if not (isinstance(v, list)):
            key_outmaps.append(v)
        else:
            key_outmaps.extend(v)
    if set(key_outmaps) != set(jcols['col_start'].keys()):
        print("Error: Please verify col_key's output columns match col_start's keys.")
        print("\tColumn name differences:")
        print("\tcol_keys:", set(key_outmaps) - set(jcols['col_start'].keys()))
        print("\tcol_start:", set(jcols['col_start'].keys()) - set(key_outmaps))
        return False

    # Check col_start cell location format
    for v in jcols['col_start'].values():
        if not (re.match(r'^[a-zA-Z]{1,3}\d+$', v)):
            print("Error: Location string \"{}\" does not follow Excel coordinate convention.".format(v))
            return False

    return True

def read_input_excel(in_file, jcols):
    """
    Reads the in_file excel into a DataFrame, saving all data points as objects. If the keys of col_keys from jcols exist
    in the DataFrame's columns (that is, if the input column names all have matches in the input excel), then the DataFrame
    is returned. Otherwise, an empty DataFrame is returned.

    in_file - The input excel file to read and copy data from.
    jcols - The JSON dictionary of input column names, template column names, and template header locations.
    """
    # Read the excel starting from the first cell, assuming the first row is the table header and all subsequent rows
    # contain corresponding data. All data is saved as objects.
    df = pd.read_excel(in_file, dtype=object)

    # Check the input column names from col_keys exist in the DataFrame columns (the excel table header)
    if set(jcols['col_keys'].keys()).issubset(set(df.columns)):
        return df
    else:
        print("Error: Please verify the columns in the column configuration file exist in the input Excel sheet.")
        print("\tcol_keys:", set(jcols['col_keys'].keys()) - set(df.columns))
        return pd.DataFrame()

def get_rename_cols(jcols):
    """
    Returns a dictionary of only the column mappings that have a one-to-one mapping. That is, one string key maps to one
    string value. Since in almost every case we are not modifying any data, only the column names that the data will be
    stored under, the headers will be modified to match the template headers.

    jcols - The JSON dictionary of input column names, template column names, and template header locations.
    """
    renames = {}
    for k,v in jcols['col_keys'].items():
        if (not (isinstance(v, list))) & (isinstance(v, str)):
            renames[k] = v
    return renames

def get_name_cols(jcols, renames):
    """
    Validates the name-splitting column and its mapping. It collects the remaining column mappings from col_keys that were
    lists or not strings, and checks that there is only one such entity. Then, it checks that there are two elements in the
    mapped list. Then, it checks that the input column name contains "name" and that the template columns are "last name"
    and "first name", ignoring capitalization. If all of these checks are passed, it will return the input column name that
    needs to be split into the two Last Name and First Name columns.

    jcols - The JSON dictionary of input column names, template column names, and template header locations.
    renames - The subset of column names from col_keys from jcols with one-to-one string mappings.
    """
    list_key = list(set(jcols['col_keys']) - set(renames.keys()))

    # Only allowed one mapping of input column to a list of output columns
    if len(list_key) > 1:
        print("Error: More than one key from col_key was found to map to a list.")
        print("\t", list_key)
        return ""
    elif len(list_key) < 1:
        return ""

    name_in_col = list_key[0]
    name_out_cols = jcols['col_keys'][name_in_col]
    # Only two output columns allowed
    if len(name_out_cols) != 2:
        print("Error: Two output columns are required for name splitting. Found {}.".format(len(name_out_cols)))
        print("\t {}: {}".format(name_in_col, name_out_cols))
        return ""

    # Ensure it is a Name column mapping to "Last Name" and "First Name" columns
    valid_in = False
    valid_last_out = False
    valid_first_out = False
    if "name" in name_in_col.lower():
        valid_in = True
    else:
        print("Error: {} is not valid for name-splitting.".format(name_in_col))
        print("\t {}: {}".format(name_in_col, name_out_cols))
    for item in name_out_cols:
        if (item.lower() == "last name") & (valid_last_out == False):
            valid_last_out = True
        elif (item.lower() == "first name") & (valid_first_out == False):
            valid_first_out = True
        else:
            print("Error: Name splitting output columns do not match \"Last Name\" and \"First Name\".")
            print("\t", name_out_cols)

    if valid_in & valid_last_out & valid_first_out:
        return name_in_col
    else:
        return ""

def split_client_account_name(full_name):
    """
    Splits the string by whitespaces and returns a labeled Series with the final element as the Last Name and the first
    element as the First Name. If the full_name is not a string, it is likely to have read in an empty (nan), and so it
    will return "None" for both the Last Name and First Name.

    full_name - The string to be split into first and last names.
    """
    if isinstance(full_name, str):
        names = full_name.split()
        return pd.Series([names[-1].capitalize(), names[0].capitalize()], index=['Last Name','First Name'])
    else:
        return pd.Series(["None", "None"], index=['Last Name','First Name'])

def validate_jcols_starts(jcols, worksheet):
    """
    Validates the header name and coordinates given in the column configuration JSON match the template Excel sheet.

    jcols - The JSON dictionary of input column names, template column names, and template header locations.
    worksheet - The worksheet of the template Excel.
    """
    for k,v in jcols['col_start'].items():
        if worksheet[v].value != k:
            print("Error: {} does not match {}. Please check your template file.".format(k, worksheet[v].value))
            return False
    return True

def read_file_data(col_file, in_file):
    """
    Reads the input file data.

    col_file - The name of the column configuration JSON file.
    in_file - The input excel file to read and copy data from.
    """
    # Read JSON
    jcols = read_col_json(col_file)
    if len(jcols) >= 2:
        df = read_input_excel(in_file, jcols)
        if len(df) != 0:
            filtered_df = pd.DataFrame(df[list(jcols['col_keys'].keys())])
            return jcols, filtered_df
        else:
            raise Exception("the input Excel file could not be read.")
    else:
        raise Exception("the column configuration JSON file was invalid.")

def process_df(jcols, df):
    """
    Renames the one-to-one column mappings. If a name-splitting mapping is found, it will split the names.

    jcols - The JSON dictionary of input column names, template column names, and template header locations.
    df - The dataframe consisting of information from the original input excel file that will be copied into the new workbook.
    """
    # Finds and resolves simple (one-to-one) column mappings
    rename_cols = get_rename_cols(jcols)
    df.rename(columns=rename_cols, inplace=True)

    # If there are leftover columns that were not simple mappings, attempt
    # name-splitting
    if len(rename_cols) != len(jcols['col_keys']):
        # Validate name columns
        name_in_col = get_name_cols(jcols, rename_cols)
        if name_in_col != "":
            # If passed validation, split the client name and change df
            name_df = df[name_in_col].apply(split_client_account_name)
            df = (pd.concat([df, name_df], axis=1))
            df.drop(columns=name_in_col, inplace=True)
        else:
            # If failed validation, the app will terminate
            raise Exception("the column configuration JSON file was invalid. Check name-splitting indicators and lists.")

    return df

def write_to_workbook(template_file, out_file, jcols, df):
    """
    Loads the template file, adds the new information, and saves this changed file to a new file named out_file.

    template_file - The file providing the formatting of the output file.
    out_file - The name to use when saving the amended output file.
    """
    # Load the workbook
    wb = load_workbook(template_file)
    ws = wb.active

    # Verify the columns match in the workbook
    if validate_jcols_starts(jcols, ws):
        # Write the data into the document
        for col_name, col_loc in jcols['col_start'].items():
            row, col = cell.coordinate_to_tuple(col_loc)
            row += 1
            for d in df[col_name].values:
                ws.cell(row=row, column=col).value = d
                row += 1

    wb.save(out_file)
    wb.close()

def main(*args):
    # Parse cmd arg input
    in_file = ""
    template_file = ""
    out_file = ""
    col_file = ""
    if len(args) == 4:
        in_file = args[0]
        template_file = args[1]
        out_file = args[2]
        col_file = args[3]
    elif len(args) == 0:
        in_file, template_file, out_file, col_file = get_default_files()
        print("Using default file inputs.")
        print("\tinput: {}\n\ttemplate: {}\n\toutput: {}\n\tcolumns: {}".format(
            in_file, template_file, out_file, col_file
        ))
    else:
        e_string = """the correct number of arguments was not passed.
        Please check that your input matches
        `python pks.py <input.xlsx> <template.xlsx> <output.xlsx> <col_config.json>`
        for command line style input of file names,
        or `python pks.py` for direct modification of variables in the get_default_files function.
        """
        raise Exception(e_string)

    if (out_file == in_file) | (out_file == template_file) | (out_file == col_file):
        print("Error: Given more than one file sharing the same name.")
        raise Exception("the output file has the same name as another file. Please use a different name for the output file.")

    jcols, df = read_file_data(col_file, in_file)
    df = process_df(jcols, df)
    write_to_workbook(template_file, out_file, jcols, df)

if __name__ == "__main__":
    main(*sys.argv[1:])
